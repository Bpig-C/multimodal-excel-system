"""品质失效案例处理器 - 管线A骨架（真正的AI标注走管线B的ExcelProcessingService）"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from .base import BaseProcessor, ExportConfig, GraphConfig


CANONICAL_SECOND_FIELD = '客户/发生工程/供应商'
SECOND_FIELD_ALIASES = [
    '客户/发生工程/供应商',
    '问题来源',
    '供应商',
    '客户',
]


def _normalize_header_name(name: Any) -> str:
    """统一表头比较口径：去首尾空白并移除中间空白。"""
    if name is None:
        return ''
    text = str(name).strip()
    return ''.join(text.split())


NORMALIZED_SECOND_FIELD_ALIASES = {_normalize_header_name(c) for c in SECOND_FIELD_ALIASES}


# 源字段名 → 标准实体字段名
FIELD_MAPPING = {
    'classification': '问题分类',
    'customer': '客户/发生工程/供应商',
    'defect': '质量问题',
    'description': '问题描述',
    'handling': '问题处理',
    'root_cause': '原因分析',
    'measures': '采取措施',
    'images': '图片',
}

# 缺失字段默认值
DEFAULT_VALUES = {
    '问题分类': '未分类',
    '客户/发生工程/供应商': '未知',
    '质量问题': '未分类',
    '问题描述': '',
    '问题处理': '',
    '原因分析': '',
    '采取措施': '',
}

# Excel 列名重映射（别名 → 标准名）
COLUMN_MAPPINGS = {
    '问题来源': CANONICAL_SECOND_FIELD,
    '供应商': CANONICAL_SECOND_FIELD,
    '客户': CANONICAL_SECOND_FIELD,
}

# 实体标注字段 [(源字段名, 实体标签)]
ENTITY_FIELDS = [
    ('问题分类', 'DEFECT_CATEGORY'),
    ('客户/发生工程/供应商', 'CUSTOMER'),
    ('质量问题', 'QUALITY_ISSUE'),
    ('问题描述', 'DESCRIPTION'),
    ('问题处理', 'HANDLING'),
    ('原因分析', 'ROOT_CAUSE'),
    ('采取措施', 'MEASURES'),
]

# 关系映射 [(源字段名, 关系类型)]
RELATION_MAPPINGS = [
    ('客户/发生工程/供应商', 'hasCustomer'),
    ('质量问题', 'hasDefect'),
    ('原因分析', 'hasRootCause'),
    ('采取措施', 'hasMeasure'),
    ('问题处理', 'hasHandling'),
]

# Q&A 查询模板
QA_QUERY = "这个品质失效案例的问题分类是什么？原因分析是什么？采取了什么措施？"

# CLIP 导出字段顺序
CLIP_FIELDS = [
    '问题分类', '客户/发生工程/供应商', '质量问题',
    '问题描述', '问题处理', '原因分析', '采取措施',
]


class FailureCaseProcessor(BaseProcessor):
    """品质失效案例处理器（管线A骨架）"""

    @property
    def name(self) -> str:
        return 'failure_case'

    @property
    def display_name(self) -> str:
        return '品质失效案例'

    @property
    def field_mapping(self):
        return FIELD_MAPPING

    @property
    def default_values(self):
        return DEFAULT_VALUES

    @property
    def column_mappings(self):
        return COLUMN_MAPPINGS

    def parse_excel(self, file_path: str, output_dir: str) -> dict:
        """
        解析品质失效案例 Excel，返回基本信息。
        output_dir 由 DocumentProcessor 传入用于写 JSON，骨架实现不使用。
        深度处理由管线B的 ExcelProcessingService 负责。
        """
        REQUIRED_COLUMNS = [
            '问题分类', '质量问题', '问题描述', '问题处理', '原因分析', '采取措施',
        ]
        normalized_required = {_normalize_header_name(col) for col in REQUIRED_COLUMNS}

        file_path = Path(file_path)
        wb = openpyxl.load_workbook(file_path, data_only=True)

        all_tables: list[dict] = []
        found_header = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # 找到表头行（含所有必要列的行）
            header = None
            header_row_idx = None
            for i, row in enumerate(rows):
                row_strs = [str(c).strip() if c is not None else "" for c in row]
                normalized_row = {_normalize_header_name(c) for c in row_strs if c}
                has_required = normalized_required.issubset(normalized_row)
                has_second_field = bool(normalized_row & NORMALIZED_SECOND_FIELD_ALIASES)

                if has_required and has_second_field:
                    header_row_idx = i
                    header = [
                        CANONICAL_SECOND_FIELD
                        if _normalize_header_name(c) in NORMALIZED_SECOND_FIELD_ALIASES
                        else c
                        for c in row_strs
                    ]
                    found_header = True
                    break

            if header is None:
                continue

            # 统计数据行数（跳过空行）
            record_count = 0
            for row in rows[header_row_idx + 1:]:
                if any(c is not None and str(c).strip() for c in row):
                    record_count += 1

            all_tables.append({
                "sheet_name": sheet_name,
                "record_count": record_count,
                "columns": [h for h in header if h],
            })

        wb.close()

        if not found_header:
            raise ValueError(
                f"文件格式不符：品质失效案例 Excel 必须包含以下所有列：{REQUIRED_COLUMNS + [CANONICAL_SECOND_FIELD]}，"
                f"其中“{CANONICAL_SECOND_FIELD}”支持别名 {SECOND_FIELD_ALIASES}，"
                f"请确认上传了正确的文件。"
            )

        return {
            "data_source": file_path.stem,
            "table_count": len(all_tables),
            "tables": all_tables,
            "total_records": sum(t["record_count"] for t in all_tables),
            "image_directory": str(Path(output_dir) / "images"),
            "status": "success",
        }

    def get_export_config(self) -> ExportConfig:
        return ExportConfig(
            field_names={v: v for v in FIELD_MAPPING.values()},
            entity_fields=ENTITY_FIELDS,
            relation_mappings=RELATION_MAPPINGS,
            qa_query=QA_QUERY,
            clip_fields=CLIP_FIELDS,
            event_id_field='问题分类',
        )

    def get_graph_config(self) -> GraphConfig:
        """品质案例不使用知识图谱（管线B负责实体关系标注）"""
        return GraphConfig(
            node_types={},
            edge_labels={},
        )

    def build_entity_text(self, record: dict, data_source: str = None) -> str:
        """
        管线A骨架：品质案例的语料生成由管线B负责，
        此处返回简单的纯文本拼接，仅用于兼容接口。
        """
        parts = []
        for cn_key, en_key in FIELD_MAPPING.items():
            val = record.get(en_key) or record.get(cn_key, "")
            if val and str(val).strip():
                parts.append(f"{cn_key}：{val}")
        return "\n".join(parts)
